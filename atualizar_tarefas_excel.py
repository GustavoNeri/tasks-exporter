# -*- coding: utf-8 -*-
import os
import re
import sys
import math
import shutil
import datetime
try:
    import urllib.parse as urlparse
except ImportError:
    import urlparse
import requests
from bs4 import BeautifulSoup
import openpyxl
from docx import Document
from docx.shared import Pt, RGBColor
from copy import copy
from copy import deepcopy

# ==========================================
# CONFIGURAÇÕES DO USUÁRIO
# ==========================================
USUARIO = "gustavo.souza@amlconsulting.com.br"
SENHA = "Gu@aml"
IDS_FUNCIONARIOS = [65, 270, 271, 74, 75, 68]  # Adicione aqui os IDs dos funcionários (ex: [278, 279])
# 65  - Diovane Barbieri Gabriel
# 270 - Gustavo Neri
# 271 - Marlon David Domingos
# 74  - Sebastião Vitor dos Santos
# 75  - Thallys Henrique Lima da Silva
# 68  - Vítor Fontanari da Silva
# [65, 270, 271, 74, 75, 68]

# Mapa de ID para nome e cargo do funcionário
FUNCIONARIOS_MAP = {
    65:  {"nome": "Diovane Barbieri Gabriel", "cargo": "Analista de Implantação"},
    270: {"nome": "Gustavo Neri", "cargo": "Tech Lead e Coordenador"},
    271: {"nome": "Marlon David Domingos", "cargo": "Analista de Desenvolvimento"},
    74:  {"nome": "Sebastião Vitor dos Santos", "cargo": "Analista de Desenvolvimento"},
    75:  {"nome": "Thallys Henrique Lima da Silva", "cargo": "Analista de Desenvolvimento"},
    68:  {"nome": "Vítor Fontanari da Silva", "cargo": "Analista de Implantação"}
}
# ==========================================

# Datas da planilha consolidada
def obter_periodo_datas():
    excel_inicio = datetime.date(2026, 5, 1)
    excel_fim = datetime.date(2026, 12, 31)
    excel_periodo_de_str = excel_inicio.strftime("%d/%m/%Y")
    excel_periodo_ate_str = excel_fim.strftime("%d/%m/%Y")
    return excel_inicio, excel_fim, excel_periodo_de_str, excel_periodo_ate_str

def parse_date(date_str):
    if not date_str:
        return None
    date_str = date_str.strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None

def subtrair_dias_uteis(data_referencia, dias_para_subtrair):
    data_atual = data_referencia
    dias_subtraidos = 0
    while dias_subtraidos < dias_para_subtrair:
        data_atual -= datetime.timedelta(days=1)
        if data_atual.weekday() < 5:  # 0 a 4 representam Segunda a Sexta
            dias_subtraidos += 1
    return data_atual

def encontrar_input(soup, selector, fallback_type=None, fallback_names=None):
    # Tenta usar o seletor exato
    elements = soup.select(selector)
    if elements:
        return elements[0]
        
    # Se falhar e contiver 'tbody', tenta sem o 'tbody' (parser do bs4 às vezes omite)
    if "tbody" in selector:
        selector_no_tbody = selector.replace("> tbody >", ">").replace("tbody >", "")
        elements = soup.select(selector_no_tbody)
        if elements:
            return elements[0]
            
    # Fallback por tipo de input
    if fallback_type:
        elements = soup.find_all("input", type=fallback_type)
        if elements:
            return elements[0]
            
    # Fallback por nome do atributo
    if fallback_names:
        for name in fallback_names:
            element = soup.find("input", attrs={"name": name})
            if element:
                return element
    return None

def normalizar_texto(texto):
    if not texto:
        return ""
    import unicodedata
    texto_str = str(texto).strip().rstrip(":").strip()
    nfkd_form = unicodedata.normalize('NFKD', texto_str)
    return "".join([c for c in nfkd_form if not unicodedata.combining(c)]).lower()

def encontrar_links_por_coluna(soup, nome_coluna="titulo"):
    links = []
    nome_coluna_norm = normalizar_texto(nome_coluna)
    for table in soup.find_all("table"):
        first_row = table.find("tr")
        if not first_row:
            continue
            
        # Pega apenas as células do cabeçalho (primeira linha da tabela)
        headers = [normalizar_texto(cell.get_text(strip=True)) for cell in first_row.find_all(["th", "td"])]
        
        col_idx = -1
        for idx, h in enumerate(headers):
            if nome_coluna_norm in h or h == nome_coluna_norm:
                col_idx = idx
                break
        
        if col_idx != -1:
            # Encontrou a coluna. Agora itera pelas linhas de dados (pulando o cabeçalho)
            for row in table.find_all("tr")[1:]:
                cells = row.find_all("td")
                if len(cells) > col_idx:
                    cell = cells[col_idx]
                    a = cell.find("a", href=True)
                    if a:
                        links.append(a["href"])
    return list(dict.fromkeys(links))

def extrair_links_tarefas(soup):
    task_links = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        # Padrões relaxados para corresponder a /atividade/ver, /tarefa/visualizar, etc.
        patterns = [
            "/atividade/ver", "/atividade/visualizar", "/atividade/detalhe", "/atividade/exibir", "id_atividade=",
            "/tarefa/ver", "/tarefa/visualizar", "/tarefa/detalhe", "/tarefa/exibir", "id_tarefa="
        ]
        if any(pat in href for pat in patterns):
            if "javascript:" not in href and "excluir" not in href and "deletar" not in href:
                task_links.append(href)
    return list(dict.fromkeys(task_links))


def extrair_texto_apos_dois_pontos(text):
    """Extrai apenas o texto após os dois pontos"""
    if ":" not in text:
        return text
    parts = text.split(":", 1)
    return parts[1].strip()

def obter_valor_select(soup, nome_campo):
    """
    Retorna o texto da opção selecionada de um SELECT.
    """

    select = soup.find("select", {"name": nome_campo})

    if not select:
        return ""

    for option in select.find_all("option"):
        if option.has_attr("selected"):
            return option.get_text(" ", strip=True)

    return ""

def extrair_dados_tarefa_por_seletores(soup):
    """Extrai dados da tarefa usando seletores corretos para a estrutura real"""
    dados = {}
    
    # Campos atuais dos selects
    status_atual = obter_valor_select(
        soup,
        "id_status"
    )
    if status_atual:
        dados["status"] = status_atual

    responsavel_atual = obter_valor_select(
        soup,
        "id_usuario_resp"
    )
    if responsavel_atual:
        dados["responsavel"] = responsavel_atual

    usuario_resp = obter_valor_select(
        soup,
        "id_usuario_para"
    )
    if usuario_resp:
        dados["incluido_por"] = usuario_resp

    prioridade = obter_valor_select(
        soup,
        "id_prioridade"
    )
    if prioridade:
        dados["prioridade"] = prioridade

    try:
        horas_input = soup.find("input", {"name": "horas_estimadas"})

        if horas_input:
            horas = horas_input.get("value", "").strip()

            if horas:
                dados["estimativa_raw"] = horas

    except Exception:
        pass

    try:
        # Empresa - da tabela barraTarefa, primeira coluna (td[0])
        empresa_elem = soup.select("table.barraTarefa tr td")
        if empresa_elem and len(empresa_elem) > 0:
            empresa_span = empresa_elem[0].find("span")
            if empresa_span:
                dados["empresa"] = empresa_span.get_text(strip=True)
    except Exception as e:
        pass
    
    try:
        # Projeto - da tabela barraTarefa, terceira coluna (td[2])
        projeto_elem = soup.select("table.barraTarefa tr td")
        if projeto_elem and len(projeto_elem) > 2:
            projeto_span = projeto_elem[2].find("span")
            if projeto_span:
                dados["projeto"] = projeto_span.get_text(strip=True)
    except Exception as e:
        pass
    
    try:
        # Atividade/Titulo - procura por <strong> que vem após "Título:"
        titulo_div = None
        for div in soup.find_all("div", class_="titulo"):
            if "Título" in div.get_text():
                titulo_div = div
                break
        
        if titulo_div:
            strong_elem = titulo_div.find_next("strong")
            if strong_elem:
                dados["atividade"] = strong_elem.get_text(strip=True)
        else:
            titulo_elem = soup.find("strong", string=lambda x: x and len(x) > 20 and "Empresa" not in x)
            if titulo_elem:
                dados["atividade"] = titulo_elem.get_text(strip=True)
    except Exception as e:
        pass
    
    try:
        # Bloco de informacoes - procura por ul class="info-container"
        info_ul = soup.find("ul", class_="info-container")
        
        if info_ul:
            lis = info_ul.find_all("li")
                        
            if len(lis) > 2:
                incluido_em_text = lis[2].get_text(strip=True)
                match_data = re.search(r"\d{1,2}/\d{1,2}/\d{4}", incluido_em_text)
                if match_data:
                    dados["incluido_em"] = match_data.group(0)
            
            if len(lis) > 7:
                entrega_em_text = lis[7].get_text(strip=True)
                match_data = re.search(r"\d{1,2}/\d{1,2}/\d{4}", entrega_em_text)
                if match_data:
                    dados["entrega_em"] = match_data.group(0)
            
            if len(lis) > 8:
                ordem_text = lis[8].get_text(strip=True)
                dados["ordem"] = extrair_texto_apos_dois_pontos(ordem_text)

    except Exception as e:
        print("  ! Erro ao extrair informacoes: " + str(e))
        import traceback
        traceback.print_exc()
    
    # Extrair descrição da tarefa
    try:
        descricao_tarefa = extrair_descricao_tarefa(soup)
        if descricao_tarefa:
            dados["descricao_tarefa"] = descricao_tarefa
        else:
            dados["descricao_tarefa"] = ""
    except Exception as e:
        dados["descricao_tarefa"] = ""
    
    return dados

def preencher_linha(sheet, row_idx, dados):
    aplicar_formatacao_linha(sheet, row_idx)
    sheet.cell(row=row_idx, column=1, value=dados.get("empresa"))
    sheet.cell(row=row_idx, column=2, value=dados.get("incluido_em"))
    sheet.cell(row=row_idx, column=3, value=dados.get("incluido_por"))
    sheet.cell(row=row_idx, column=4, value=dados.get("responsavel"))
    sheet.cell(row=row_idx, column=5, value=dados.get("projeto"))
    sheet.cell(row=row_idx, column=6, value=dados.get("atividade"))
    sheet.cell(row=row_idx, column=7, value=dados.get("ordem"))
    sheet.cell(row=row_idx, column=8, value=dados.get("comecar_em"))
    sheet.cell(row=row_idx, column=9, value=dados.get("entrega_em"))
    sheet.cell(row=row_idx, column=10, value=dados.get("estimativa_horas"))
    sheet.cell(row=row_idx, column=11, value=dados.get("status"))
    sheet.cell(row=row_idx, column=12, value="")  # Observação (Em branco)

def aplicar_formatacao_linha(sheet, row_idx, template_row=2):
    """Mantem a formatacao da linha modelo ao inserir novas linhas."""
    if row_idx == template_row:
        return
    for col_idx in range(1, sheet.max_column + 1):
        origem = sheet.cell(row=template_row, column=col_idx)
        destino = sheet.cell(row=row_idx, column=col_idx)
        if origem.has_style:
            destino._style = copy(origem._style)
        if origem.number_format:
            destino.number_format = origem.number_format
        if origem.alignment:
            destino.alignment = copy(origem.alignment)
        if origem.font:
            destino.font = copy(origem.font)
        if origem.fill:
            destino.fill = copy(origem.fill)
        if origem.border:
            destino.border = copy(origem.border)
        if origem.protection:
            destino.protection = copy(origem.protection)

def limpar_linhas_dados(sheet, start_row=2):
    for row in range(start_row, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None

def preparar_planilha_excel(template_path, output_path):
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    sheet_lista = wb.active
    sheet_lista.title = "Lista de Tarefas"
    limpar_linhas_dados(sheet_lista)
    sheet_realizado = wb.copy_worksheet(sheet_lista)
    sheet_realizado.title = "Realizado"
    return wb, sheet_lista, sheet_realizado

def status_eh_fechado(status):
    return normalizar_texto(status) == "fechado"

def main():
    if USUARIO == "SEU_USUARIO" or SENHA == "SUA_SENHA":
        print("Erro: Por favor, configure seu USUÁRIO e SENHA no início do script.")
        sys.exit(1)
    
    # Obter as datas
    excel_inicio, excel_fim, excel_periodo_de_str, excel_periodo_ate_str = obter_periodo_datas()

    print("Processando planilha de " + excel_periodo_de_str + " a " + excel_periodo_ate_str)
    
    # Verificar se o modelo Excel existe
    template_name = "00-Prestador-Mapeamento-de-Tarefas-Semana-XX-XX-XXXX-a-XX-XX-XXXX.xlsx"
    workspace_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(workspace_dir, template_name)
    
    if not os.path.exists(template_path):
        print("ERRO: Modelo Excel nao encontrado: " + template_name)
        sys.exit(1)
    
    # Iniciar sessão
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.0.0 Safari/537.36"
    })
    
    # 1. Login disparado pelo redirect de acesso à lista
    primeiro_id = IDS_FUNCIONARIOS[0] if IDS_FUNCIONARIOS else 278
    list_redirect_url = "https://tasks.amlgroup.com.br/atividade/lista?id_usuario_responsavel=" + str(primeiro_id) + "&periodoDe=" + excel_periodo_de_str + "&periodoAte=" + excel_periodo_ate_str
    try:
        response = session.get(list_redirect_url)
    except Exception as e:
        print("ERRO ao conectar ao portal: " + str(e))
        sys.exit(1)
        
    soup = BeautifulSoup(response.text, "html.parser")
    form = soup.find("form", id="form1") or soup.find("form")
    
    form_data = {}
    if form:
        for ipt in form.find_all("input"):
            name = ipt.get("name")
            if name:
                form_data[name] = ipt.get("value", "")
        for btn in form.find_all("input", type="submit"):
            name = btn.get("name")
            if name:
                form_data[name] = btn.get("value", "")
                
    user_selector = "#form1 > table > tbody > tr:nth-child(2) > td > table > tbody > tr:nth-child(1) > td:nth-child(2) > input"
    pass_selector = "#form1 > table > tbody > tr:nth-child(2) > td > table > tbody > tr:nth-child(2) > td:nth-child(2) > input"
    
    user_input = encontrar_input(soup, user_selector, fallback_type="text", fallback_names=["usuario", "login", "email"])
    pass_input = encontrar_input(soup, pass_selector, fallback_type="password", fallback_names=["senha", "password"])
    
    if user_input:
        user_name_attr = user_input.get("name")
        if user_name_attr:
            form_data[user_name_attr] = USUARIO
    if pass_input:
        pass_name_attr = pass_input.get("name")
        if pass_name_attr:
            form_data[pass_name_attr] = SENHA
            
    action_url = form.get("action", "") if form else ""
    submit_url = urlparse.urljoin(response.url, action_url) if action_url else response.url
    
    try:
        login_resp = session.post(submit_url, data=form_data)
    except Exception as e:
        print("ERRO ao fazer login: " + str(e))
        sys.exit(1)
        
    if "senha" in login_resp.text.lower() and "login" in login_resp.text.lower() and login_resp.url == response.url:
        print("ERRO: Autenticacao falhou")
        sys.exit(1)
    
    excel_output_name = "Mapeamento-de-Tarefas-Tecnologia-AML-Monitor-Desenvolvimento-2026.xlsx"
    excel_output_path = os.path.join(workspace_dir, excel_output_name)
    wb_excel, sheet_lista, sheet_realizado = preparar_planilha_excel(template_path, excel_output_path)
    excel_linha_lista = 2
    excel_linha_realizado = 2
    
    total_inserido = 0
    total_excel_lista = 0
    total_excel_realizado = 0
    tarefas_por_funcionario = {}  # { emp_id: [lista de tarefas] }
    
    for emp_id in IDS_FUNCIONARIOS:
        tarefas_por_funcionario[emp_id] = []
        
        # Obter dados do funcionário
        func_info = FUNCIONARIOS_MAP.get(emp_id, {"nome": "Funcionario"})
        
        print("Extraindo planilha: " + func_info["nome"])
        list_url = "https://tasks.amlgroup.com.br/atividade/lista"
        params = {
            "id_empresa_grupo": "",
            "id_empresa_area": "",
            "id_usuario_responsavel": str(emp_id),
            "id_projeto": "",
            "status": "Todos",
            "periodoDe": excel_periodo_de_str,
            "periodoAte": excel_periodo_ate_str,
            "orderby": "tarefa_informacao.data_entrega, tarefa_informacao.ordem ASC"
        }
        
        try:
            list_resp = session.get(list_url, params=params)
        except Exception as e:
            print("  ERRO ao buscar tarefas da planilha: " + str(e))
            continue
            
        if list_resp.status_code != 200:
            print("  ERRO planilha: status " + str(list_resp.status_code))
            continue
            
        list_soup = BeautifulSoup(list_resp.text, "html.parser")
        
        # Tenta extrair links das tarefas
        task_links = encontrar_links_por_coluna(list_soup, "titulo")
        if not task_links:
            task_links = encontrar_links_por_coluna(list_soup, "título")
        if not task_links:
            task_links = encontrar_links_por_coluna(list_soup, "atividade")
        if not task_links:
            task_links = extrair_links_tarefas(list_soup)
            
        if not task_links:
            continue
        total_inserido_func = 0
        
        for link in task_links:
            task_url = urlparse.urljoin(list_resp.url, link)
            try:
                task_resp = session.get(task_url)
            except Exception as e:
                print("  ERRO ao carregar tarefa: " + str(e))
                continue
                
            if task_resp.status_code != 200:
                continue
                
            task_soup = BeautifulSoup(task_resp.text, "html.parser")
            
            debug_file = "debug_page.html"
            if not os.path.exists(debug_file):
                with open(debug_file, "w", encoding="utf-8") as f:
                    f.write(task_resp.text)
            
            dados_tarefa = extrair_dados_tarefa_por_seletores(task_soup)

            
            # Obter valores extraídos
            empresa = dados_tarefa.get("empresa", "")
            incluido_em = dados_tarefa.get("incluido_em", "")
            incluido_por = dados_tarefa.get("incluido_por", "")
            responsavel = dados_tarefa.get("responsavel", "")
            projeto = dados_tarefa.get("projeto", "")
            atividade = dados_tarefa.get("atividade", "")
            ordem = dados_tarefa.get("ordem", "")
            entrega_em_str = dados_tarefa.get("entrega_em", "")
            estimativa_raw = dados_tarefa.get("estimativa_raw", "0")
            status = dados_tarefa.get("status", "")
            
            # Tratamento de datas
            entrega_em_date = parse_date(entrega_em_str)
            if entrega_em_date:
                entrega_em_str = entrega_em_date.strftime("%d/%m/%Y")
                
            incluido_em_date = parse_date(incluido_em)
            if incluido_em_date:
                incluido_em = incluido_em_date.strftime("%d/%m/%Y")
            
            # Tratamento de horas
            horas = float(estimativa_raw) if estimativa_raw else 0.0
            
            # Calcular "Começar em" = Entrega em - (estimativa / 8 dias úteis)
            comecar_em_str = ""
            if entrega_em_date and horas > 0:
                # 8 horas por dia útil
                dias_uteis = math.ceil(horas / 8.0)
                comecar_em_date = subtrair_dias_uteis(entrega_em_date, dias_uteis)
                comecar_em_str = comecar_em_date.strftime("%d/%m/%Y")
            
            estimativa = int(horas) if horas.is_integer() else horas
            
            dados_linha = {
                "empresa": empresa,
                "incluido_em": incluido_em,
                "incluido_por": incluido_por,
                "responsavel": responsavel,
                "projeto": projeto,
                "atividade": atividade,
                "ordem": ordem,
                "comecar_em": comecar_em_str,
                "entrega_em": entrega_em_str,
                "estimativa_horas": estimativa,
                "status": status
            }
            
            if status_eh_fechado(status):
                preencher_linha(sheet_realizado, excel_linha_realizado, dados_linha)
                excel_linha_realizado += 1
                total_excel_realizado += 1
            else:
                preencher_linha(sheet_lista, excel_linha_lista, dados_linha)
                excel_linha_lista += 1
                total_excel_lista += 1
            
            total_inserido_func += 1
            total_inserido += 1
        
        print("  Tarefas encontradas: " + str(total_inserido_func))
    
    if total_excel_lista > 0 or total_excel_realizado > 0:
        wb_excel.save(excel_output_path)
        wb_excel.close()
        print("Excel consolidado: " + excel_output_name)
        print("  Lista de Tarefas: " + str(total_excel_lista))
        print("  Realizado: " + str(total_excel_realizado))
    else:
        wb_excel.close()
        if os.path.exists(excel_output_path):
            os.remove(excel_output_path)
        print("Excel consolidado nao gerado: nenhuma tarefa encontrada")
        
    # Limpar arquivo de debug ao finalizar o processo
    debug_file = "debug_page.html"
    if os.path.exists(debug_file):
        try:
            os.remove(debug_file)
        except Exception as e:
            print("ERRO ao remover arquivo de debug: " + str(e))
    
    print("\nProcessamento concluido!")

if __name__ == "__main__":
    main()
